#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
f = input("Enter the path to the rank file: ")
f1 = input("Enter the path to the route file: ")
rank = pd.read_excel(f, sheet_name='Rank',index_col=None, header=None)
rank = rank.iloc[1:, 1:].values


# In[2]:


print(rank)


# In[3]:


jobs = len(rank)
operations = len(rank[0])
print(jobs)
print(operations)
tmp = [[] for _ in range(jobs * operations)]
print(tmp)


# In[4]:


for i in range(len(rank)):
    for j in range(len(rank[0])):
        tmp[rank[i][j]-1] = [i+1, j+1]

print(tmp)
        


# In[5]:


route = pd.read_excel(f1, sheet_name='Sheet1',index_col=None, header=None)
#print(route)
route_array = route.iloc[1:, 1:].values
#print(route_array)
machine = [[] for i in range(operations)]
print(machine)
for i in tmp:
    #print(i)
    #print(route_array[i[1]-1][3*(i[0]-1)])
    machine[int(route_array[i[1]-1][3*(i[0]-1)].split('_')[1]) - 1].append(i[0])

print(machine)


# In[6]:


import pandas as pd
from openpyxl import load_workbook


# Create column names
col_names = ['Seq' + str(i) for i in range(1, len(machine[0])+1)]

# Create row names
row_names = ['Machine' + str(i) for i in range(1, len(machine)+1)]

# Convert the 2D array into pandas DataFrame and specify column names and row names
df = pd.DataFrame(machine, columns=col_names, index=row_names)

# Use openpyxl to load an existing Excel file
book = load_workbook(f)

# If the worksheet exists, use it; otherwise, create it
if 'Sequence' in book.sheetnames:
    writer_sheets = book['Sequence']
else:
    writer_sheets = book.create_sheet('Sequence')

# Write DataFrame into Excel worksheet
for index, row in df.iterrows():
    writer_sheets.append(list(row))

# Save file
book.save(f)


# In[ ]:




